#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Add Xofigo prescription-drug quantity from downloaded NDB Open Data.

The base app stores medical procedure counts in data/dashboard_data.json.
Xofigo is published in the prescription-drug injection quantity tables, so this
post-processor reads the downloaded raw NDB files under the repository root's
NDB-Open/20260724 folder and inserts the extracted prefecture-level series as a
regular selectable NDB item.

For editions with public-expense split, this script uses:
  - 第10回(FY2023): 処方薬_公費レセプト含まない（一括）_001499737.zip
  - 第11回(FY2024): 処方薬_公費レセプト含まない（一括）_001711930.zip
"""

from __future__ import annotations

import csv
import io
import json
import os
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
PROJECT_ROOT = APP_DIR.parents[1]
NDB_SOURCE_DIR = PROJECT_ROOT / "NDB-Open" / "20260724"
DASHBOARD_JSON = APP_DIR / "data" / "dashboard_data.json"
COUNTS_CSV = APP_DIR / "data" / "xofigo_drug_pref_counts.csv"
SOURCES_CSV = APP_DIR / "data" / "xofigo_drug_sources.csv"

DRUG_ID = "622489201"
DRUG_YJ_CODE = "4291432A1025"
DRUG_NAME = "ゾーフィゴ静注"
CLASS_CODE = "RX"
CLASS_NAME = "処方薬（注射・薬効分類別数量）"
ITEM_NAME = "ゾーフィゴ静注（薬剤数量・回分）"

PREF_NAMES = [
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
    "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
    "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県",
    "岐阜県", "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府",
    "兵庫県", "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県",
    "広島県", "山口県", "徳島県", "香川県", "愛媛県", "高知県",
    "福岡県", "佐賀県", "長崎県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
]
PREF_CODES = [f"{i:02d}" for i in range(1, 48)]
YEARS = list(range(2014, 2025))
HYPHENS = {"-", "‐", "－", "―", "‑"}


@dataclass
class SourceDef:
    edition: int
    fy_year: int
    path: Path
    kind: str
    kohi: str
    url: str


@dataclass
class ParsedRow:
    sheet: str
    row_no: int
    drug_code: str
    drug_name: str
    unit: str
    yj_code: str
    total: float | None
    pref_values: list[float | None]


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if text == "" or text in HYPHENS:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def decode_zip_name(info: zipfile.ZipInfo) -> str:
    name = info.filename
    if info.flag_bits & 0x800:
        return name
    try:
        return name.encode("cp437").decode("cp932")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name


def load_inventory() -> list[dict[str, Any]]:
    matches = list(NDB_SOURCE_DIR.glob("*/inventory.jsonl"))
    if len(matches) != 1:
        raise FileNotFoundError(f"inventory.jsonl not found uniquely under {NDB_SOURCE_DIR}")
    rows: list[dict[str, Any]] = []
    with matches[0].open(encoding="utf-8") as handle:
        for line in handle:
            if line.strip():
                rows.append(json.loads(line))
    return rows


def is_injection_pref_quantity(row: dict[str, Any]) -> bool:
    link = norm(row.get("link_name"))
    h4 = norm(row.get("h4"))
    filename = norm(row.get("filename"))
    is_injection = (
        h4 == "注射"
        or link.startswith("注射")
        or "_注射_" in filename
        or filename.startswith("薬剤_注射")
    )
    return (
        is_injection
        and ("都道府県別" in link or "都道府県別" in filename)
        and ("薬効分類別数量" in link or "薬効分類別数量" in filename)
    )


def select_sources(inventory: list[dict[str, Any]]) -> list[SourceDef]:
    sources: list[SourceDef] = []

    for edition in range(1, 10):
        matches = [
            row for row in inventory
            if row.get("edition_no") == edition and is_injection_pref_quantity(row)
        ]
        if len(matches) != 1:
            raise RuntimeError(f"edition {edition}: expected one injection-pref quantity file, got {len(matches)}")
        row = matches[0]
        sources.append(SourceDef(
            edition=edition,
            fy_year=2013 + edition,
            path=NDB_SOURCE_DIR / row["folder"] / row["filename"],
            kind="xlsx",
            kohi=norm(row.get("kohi")),
            url=norm(row.get("url")),
        ))

    no_public_zip_ids = {10: "001499737", 11: "001711930"}
    for edition, file_id in no_public_zip_ids.items():
        matches = [
            row for row in inventory
            if row.get("edition_no") == edition and file_id in norm(row.get("filename"))
        ]
        if len(matches) != 1:
            raise RuntimeError(f"edition {edition}: expected one no-public prescription zip, got {len(matches)}")
        row = matches[0]
        kohi = norm(row.get("kohi"))
        if "含まない" not in kohi:
            raise RuntimeError(f"edition {edition}: selected ZIP is not marked no-public: {kohi}")
        sources.append(SourceDef(
            edition=edition,
            fy_year=2013 + edition,
            path=NDB_SOURCE_DIR / row["folder"] / row["filename"],
            kind="zip",
            kohi=kohi,
            url=norm(row.get("url")),
        ))

    for source in sources:
        if not source.path.exists():
            raise FileNotFoundError(source.path)
    return sources


def find_header(ws: Any) -> tuple[int, int, list[int]] | None:
    for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        values = [norm(v) for v in row]
        total_cols = [i for i, value in enumerate(values) if "総計" in value]
        pref_cols = [
            i for i, value in enumerate(values)
            if len(value) == 2 and value.isdigit() and 1 <= int(value) <= 47
        ]
        if total_cols and len(pref_cols) >= 47:
            return row_no, total_cols[0], pref_cols[:47]
    return None


def parse_workbook(source: io.BytesIO | Path) -> tuple[list[ParsedRow], list[str]]:
    wb = load_workbook(source, read_only=True, data_only=True)
    parsed: list[ParsedRow] = []
    notes: list[str] = []

    for ws in wb.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        if first:
            notes.append(str(first))

        header = find_header(ws)
        if header is None:
            continue
        header_row, total_col, pref_cols = header

        for row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            values = list(row)
            text = " ".join(norm(value) for value in values[:10])
            if DRUG_NAME not in text and DRUG_YJ_CODE not in text and DRUG_ID not in text:
                continue

            parsed.append(ParsedRow(
                sheet=ws.title,
                row_no=row_no,
                drug_code=norm(values[2] if len(values) > 2 else ""),
                drug_name=norm(values[3] if len(values) > 3 else ""),
                unit=norm(values[4] if len(values) > 4 else ""),
                yj_code=norm(values[5] if len(values) > 5 else ""),
                total=to_number(values[total_col] if total_col < len(values) else None),
                pref_values=[to_number(values[col] if col < len(values) else None) for col in pref_cols],
            ))

    return parsed, notes


def parse_source(source: SourceDef) -> tuple[list[ParsedRow], list[str], str]:
    if source.kind == "xlsx":
        parsed, notes = parse_workbook(source.path)
        return parsed, notes, source.path.name

    with zipfile.ZipFile(source.path) as zf:
        target_info = None
        target_name = ""
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = decode_zip_name(info)
            base = os.path.basename(name)
            if (
                base.startswith("【注射】")
                and "都道府県別" in base
                and "薬効分類別数量" in base
                and base.lower().endswith(".xlsx")
            ):
                target_info = info
                target_name = name
                break
        if target_info is None:
            raise RuntimeError(f"injection-pref quantity workbook not found in {source.path}")
        parsed, notes = parse_workbook(io.BytesIO(zf.read(target_info)))
        return parsed, notes, f"{source.path.name}::{target_name}"


def aggregate_rows(rows: list[ParsedRow]) -> tuple[float | None, dict[str, float | None], dict[str, str]]:
    if not rows:
        return None, {code: None for code in PREF_CODES}, {code: "not_found" for code in PREF_CODES}

    total = None
    if any(row.total is not None for row in rows):
        total = sum(row.total for row in rows if row.total is not None)

    pref_counts: dict[str, float | None] = {}
    pref_status: dict[str, str] = {}
    for idx, pref_code in enumerate(PREF_CODES):
        values = [row.pref_values[idx] for row in rows]
        nums = [value for value in values if value is not None]
        has_masked = any(value is None for value in values)
        if not nums:
            pref_counts[pref_code] = None
            pref_status[pref_code] = "masked"
        else:
            summed = sum(nums)
            if summed == 0 and has_masked:
                pref_counts[pref_code] = None
                pref_status[pref_code] = "masked_plus_zero"
            else:
                pref_counts[pref_code] = round(summed, 6)
                pref_status[pref_code] = "partial" if has_masked else "shown"

    return (round(total, 6) if total is not None else None), pref_counts, pref_status


def upsert_dashboard(pref_by_year: dict[int, dict[str, float | None]], national_by_year: dict[int, float | None]) -> None:
    with DASHBOARD_JSON.open(encoding="utf-8") as handle:
        data = json.load(handle)

    codes = data["ndb"]["codes"]
    for group in codes:
        group["procedures"] = [proc for proc in group.get("procedures", []) if proc.get("id") != DRUG_ID]

    group = next((group for group in codes if group.get("code") == CLASS_CODE), None)
    if group is None:
        group = {"code": CLASS_CODE, "name": CLASS_NAME, "procedures": []}
        codes.append(group)
    else:
        group["name"] = CLASS_NAME
    group["procedures"].append({"id": DRUG_ID, "name": ITEM_NAME, "points": None})

    pref_counts = {code: {} for code in PREF_CODES}
    for year in YEARS:
        year_text = str(year)
        for pref_code in PREF_CODES:
            pref_counts[pref_code][year_text] = pref_by_year[year][pref_code]

    national_counts = {str(year): national_by_year[year] for year in YEARS}
    data["ndb"]["counts"]["pref"][DRUG_ID] = pref_counts
    data["ndb"]["counts"].setdefault("national", {})[DRUG_ID] = national_counts
    data["ndb"].setdefault("scr", {}).setdefault("pref", {}).pop(DRUG_ID, None)

    with DASHBOARD_JSON.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, separators=(",", ":"))


def write_trace_csv(
    year_results: dict[int, dict[str, Any]],
    pref_by_year: dict[int, dict[str, float | None]],
    status_by_year: dict[int, dict[str, str]],
    national_by_year: dict[int, float | None],
) -> None:
    with COUNTS_CSV.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["fy_year", "pref_code", "pref_name", "quantity", "status", "national_total", "source"])
        for year in YEARS:
            for pref_code, pref_name in zip(PREF_CODES, PREF_NAMES):
                writer.writerow([
                    year,
                    pref_code,
                    pref_name,
                    "" if pref_by_year[year][pref_code] is None else pref_by_year[year][pref_code],
                    status_by_year[year][pref_code],
                    "" if national_by_year[year] is None else national_by_year[year],
                    year_results[year]["source_label"],
                ])

    with SOURCES_CSV.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow([
            "fy_year", "edition", "kohi", "found_rows", "national_total",
            "source", "url", "workbook_note", "found_detail",
        ])
        for year in YEARS:
            result = year_results[year]
            rows = result["rows"]
            detail = "; ".join(
                f"{row.sheet} row {row.row_no} {row.drug_code} {row.drug_name} total={row.total}"
                for row in rows
            )
            writer.writerow([
                year,
                result["source"].edition,
                result["source"].kohi,
                len(rows),
                "" if national_by_year[year] is None else national_by_year[year],
                result["source_label"],
                result["source"].url,
                result["notes"][0] if result["notes"] else "",
                detail,
            ])


def main() -> int:
    if not NDB_SOURCE_DIR.exists():
        print(f"ERROR: downloaded NDB source folder not found: {NDB_SOURCE_DIR}", file=sys.stderr)
        return 1
    if not DASHBOARD_JSON.exists():
        print(f"ERROR: dashboard data not found: {DASHBOARD_JSON}", file=sys.stderr)
        return 1

    inventory = load_inventory()
    sources = select_sources(inventory)

    pref_by_year: dict[int, dict[str, float | None]] = {}
    status_by_year: dict[int, dict[str, str]] = {}
    national_by_year: dict[int, float | None] = {}
    year_results: dict[int, dict[str, Any]] = {}

    for source in sources:
        rows, notes, source_label = parse_source(source)
        if source.fy_year == 2024:
            joined_notes = "\n".join(notes)
            if "公費レセプトは除く" not in joined_notes:
                raise RuntimeError("FY2024 prescription workbook note does not confirm no-public-expense claims")

        national, pref_counts, pref_status = aggregate_rows(rows)
        pref_by_year[source.fy_year] = pref_counts
        status_by_year[source.fy_year] = pref_status
        national_by_year[source.fy_year] = national
        year_results[source.fy_year] = {
            "source": source,
            "source_label": source_label,
            "rows": rows,
            "notes": notes,
        }
        print(
            f"{source.fy_year}: rows={len(rows)} "
            f"national={'NA' if national is None else national} "
            f"source={source_label}"
        )

    missing = set(YEARS) - set(pref_by_year)
    if missing:
        raise RuntimeError(f"missing parsed years: {sorted(missing)}")

    upsert_dashboard(pref_by_year, national_by_year)
    write_trace_csv(year_results, pref_by_year, status_by_year, national_by_year)
    print(f"updated {DASHBOARD_JSON}")
    print(f"wrote {COUNTS_CSV}")
    print(f"wrote {SOURCES_CSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
